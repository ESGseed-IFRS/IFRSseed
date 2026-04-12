"""
Scope 3 배출계수를 Excel에서 추출하여 CSV로 저장하는 스크립트
"""
import sys
import csv
from pathlib import Path

# backend 경로 추가
backend_path = Path(__file__).parent.parent.parent
sys.path.insert(0, str(backend_path))

# UTF-8 인코딩 강제
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed.")
    print("Install: pip install openpyxl")
    sys.exit(1)

def extract_scope3_sheet(excel_file: str, output_csv: str):
    """Scope3_카테고리별 시트를 CSV로 추출"""
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    sheet_name = "Scope3_카테고리별"
    if sheet_name not in wb.sheetnames:
        print(f"[ERROR] Sheet '{sheet_name}' not found!")
        return
    
    ws = wb[sheet_name]
    print(f"[OK] Found sheet: {sheet_name}")
    print(f"[INFO] Rows: {ws.max_row}, Columns: {ws.max_column}")
    
    # CSV로 저장
    with open(output_csv, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        
        for row in range(1, ws.max_row + 1):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            writer.writerow(row_data)
    
    print(f"[OK] Exported to: {output_csv}")
    print(f"[INFO] Total {ws.max_row} rows exported")

if __name__ == "__main__":
    # 파일 경로
    project_root = Path(__file__).parent.parent.parent.parent
    excel_file = project_root / "GHG_배출계수_마스터_v2.xlsx"
    output_csv = project_root / "backend" / "SCOPE3_EMISSION_FACTORS.csv"
    
    if not excel_file.exists():
        print(f"[ERROR] File not found: {excel_file}")
        sys.exit(1)
    
    try:
        extract_scope3_sheet(str(excel_file), str(output_csv))
        print(f"\n[SUCCESS] Scope 3 emission factors extracted successfully!")
        print(f"[INFO] You can now view the file at: {output_csv}")
    except Exception as e:
        print(f"[ERROR] Failed to extract: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
