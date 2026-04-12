"""
GHG 배출계수 마스터 Excel 파일을 읽어서 내용을 확인하는 스크립트
"""
import sys
from pathlib import Path

# backend 경로 추가
backend_path = Path(__file__).parent.parent.parent
sys.path.insert(0, str(backend_path))

# UTF-8 인코딩 강제
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed.")
    print("Install: pip install openpyxl")
    sys.exit(1)

def read_excel_file(file_path: str):
    """Excel 파일을 읽어서 시트별 내용을 출력"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    print(f"[OK] Excel file loaded: {file_path}")
    print(f"[INFO] Sheet list: {wb.sheetnames}")
    print("=" * 100)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n[SHEET] {sheet_name}")
        print(f"[INFO] Rows: {ws.max_row}, Columns: {ws.max_column}")
        print("-" * 100)
        
        # 헤더 출력 (첫 번째 행)
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers.append(cell_value)
        print(f"[HEADER] {' | '.join(str(h) for h in headers)}")
        print("-" * 100)
        
        # 데이터 샘플 출력 (최대 10행)
        sample_rows = min(11, ws.max_row)  # 헤더 + 10행
        for row in range(2, sample_rows + 1):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            print(f"[ROW {row-1}] {' | '.join(row_data)}")
        
        if ws.max_row > 11:
            print(f"... (Total {ws.max_row - 1} data rows, showing first 10 only)")
        
        print("=" * 100)
        
        # Scope 3 관련 데이터 검색
        print(f"\n[SEARCH] Searching Scope 3 data in '{sheet_name}' sheet...")
        scope3_count = 0
        scope3_rows = []
        
        for row in range(2, ws.max_row + 1):
            row_text = ""
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    row_text += str(cell_value).lower()
            
            if "scope3" in row_text or "scope 3" in row_text or "cat." in row_text or "category" in row_text:
                scope3_count += 1
                if scope3_count <= 5:  # 처음 5개만 상세 출력
                    row_data = []
                    for col in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row=row, column=col).value
                        row_data.append(str(cell_value) if cell_value is not None else "")
                    scope3_rows.append((row, row_data))
        
        if scope3_count > 0:
            print(f"[FOUND] Scope 3 data found: {scope3_count} rows!")
            for row_num, row_data in scope3_rows:
                print(f"[ROW {row_num-1}] {' | '.join(row_data)}")
            if scope3_count > 5:
                print(f"... (Total {scope3_count} rows, showing first 5 only)")
        else:
            print(f"[INFO] No Scope 3 data found.")
        
        print("=" * 100)

if __name__ == "__main__":
    # Excel 파일 경로
    excel_file = Path(__file__).parent.parent.parent.parent / "GHG_배출계수_마스터_v2.xlsx"
    
    if not excel_file.exists():
        print(f"[ERROR] File not found: {excel_file}")
        sys.exit(1)
    
    try:
        read_excel_file(str(excel_file))
    except Exception as e:
        print(f"[ERROR] Failed to read Excel file: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
